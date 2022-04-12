import datetime
from datetime import timedelta, timezone
import math
from pathlib import Path
from sre_parse import SPECIAL_CHARS
from openpyxl import load_workbook
from models import Appointment, Employee, Job

BYPASS_DATE_AND_TIME_VALIDATION = True

OPENING_HOUR_TIME_DELTA = timedelta(hours=8)

START_LUNCH_TIME_DELTA = timedelta(hours=12)
END_LUNCH_TIME_DELTA = timedelta(hours=13)

CLOSING_HOUR_TIME_DELTA = timedelta(hours=17)

APPOINTMENT_LENGTH_TIME_DELTA = timedelta(minutes=30)

FAKER_LOCALE = 'en_CA'
FAKER_SEED = 34003

STRICT_POSTAL_CODE_CAPITALIZATION = False
FIX_SPACE_IN_POSTAL_CODE = True

# https://en.wikipedia.org/wiki/List_of_Special_Characters_for_Passwords
SPECIAL_CHARS_ALLOWED_IN_PASSWORD_XLSX_FILENAME = 'Wikipedia_List_of_Special_Characters_for_Passwords.xlsx'
SPECIAL_CHARS_ALLOWED_IN_PASSWORD_XLSX_FILEPATH = Path.cwd() / SPECIAL_CHARS_ALLOWED_IN_PASSWORD_XLSX_FILENAME

MEDICATION_XLSX_FILENAME = 'Clin_Calc_dot_com_the_top_200_drugs_of_2019.xlsx'
MEDICATION_XLSX_FILEPATH = Path.cwd() / MEDICATION_XLSX_FILENAME

NUM_ADMIN = 1 # less than 5
NUM_STAFF = 3 # less than 5
NUM_DOCTORS = 5 # less than 5

ADMIN_TITLE = 'admin'
STAFF_TITLE = 'staff'
DOCTOR_TITLE = 'doctor'
JOB_TITLES = [ADMIN_TITLE] * NUM_ADMIN + [STAFF_TITLE] * NUM_STAFF + [DOCTOR_TITLE] * NUM_DOCTORS

ADMIN_SPECIALTIES = ['geek', 'database admin', 'system admin', 'network admin', 'security admin']
STAFF_SPECIALTIES = ['receptionist', 'nurse', 'health worker', 'pharmacist', 'therapist', 'social worker']
DOCTOR_SPECIALTIES = ['general', 'pediatrics', 'cardiology', 'dermatology', 'neurology', 'obstetrics']
JOB_SPECIALTIES =   ADMIN_SPECIALTIES[0:NUM_ADMIN] + \
                    STAFF_SPECIALTIES[0:NUM_STAFF] + \
                    DOCTOR_SPECIALTIES[0:NUM_DOCTORS]
                    
INIT_NUM_PATIENTS = 60
NUM_PRESCRIPTIONS = INIT_NUM_PATIENTS

UNIT_NAMES = ['mg','mL','g','oz']

PRESCRIPTION_QUANTITIES = [1, 2, 3, 4, 5, 10, 20, 30, 40, 50, 100, 200, 300, 400, 500, 1000]

PROVINCES = ['AB','BC','MB','NB','NL','NS','NT','NU','ON','PE','QC','SK','YT']

def get_prescription_display_name(medication_name, quantity, unit_name):
    return f'{medication_name} ({str(quantity)} {unit_name})'

def generate_10_digit_with_2_letter_version_code_OHIP_number(faker):
    return "".join([str(faker.random_digit()) for i in range(0,10)]) + faker.random_letter().upper() + faker.random_letter().upper()

def get_special_chars_allowed_for_passwords_from_xlsx_file():
    wb = load_workbook(filename = SPECIAL_CHARS_ALLOWED_IN_PASSWORD_XLSX_FILEPATH)
    sheet = wb.active
    
    special_chars_allowed = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        if row[0].value == '':
            special_chars_allowed.append(' ') # space is missing from .xlsx file
        else:
            special_chars_allowed.append(row[0].value)
    return  special_chars_allowed

def get_medication_names_from_xlsx_file():
    
    wb = load_workbook(filename = MEDICATION_XLSX_FILEPATH)
    sheet = wb.active
    
    medication_names = []
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        medication_names.append(row[0].value.split('; ')[0])
        
    return medication_names

def get_todays_datetime_from_time_delta(time_delta):
    return datetime.datetime.combine(datetime.date.today(),datetime.time(0)) + time_delta

def get_todays_opening_datetime():
    return get_todays_datetime_from_time_delta(OPENING_HOUR_TIME_DELTA)
    
def get_todays_closing_datetime():
    return get_todays_datetime_from_time_delta(CLOSING_HOUR_TIME_DELTA)
    
def get_todays_starting_lunch_time_datetime():
    return get_todays_datetime_from_time_delta(START_LUNCH_TIME_DELTA)

def get_todays_ending_lunch_time_datetime():
    return get_todays_datetime_from_time_delta(END_LUNCH_TIME_DELTA)

def get_number_of_possible_appointments_available_per_one_doctor_per_day():
    return len(get_list_of_possible_available_appointment_datetimes_available_per_one_doctor_per_day())  
    
def get_list_of_possible_available_appointment_datetimes_available_per_one_doctor_per_day():
    available_appointment_datetimes = []
    
    current_datetime = get_todays_opening_datetime()
    while current_datetime < get_todays_closing_datetime():
        if current_datetime < get_todays_starting_lunch_time_datetime() or current_datetime >= get_todays_ending_lunch_time_datetime():
            available_appointment_datetimes.append(current_datetime)
        current_datetime += APPOINTMENT_LENGTH_TIME_DELTA 

    return available_appointment_datetimes

def get_number_of_appointments_available_today(num_doctors):
    time_available_in_day_time_delta = CLOSING_HOUR_TIME_DELTA - OPENING_HOUR_TIME_DELTA - \
                                       (END_LUNCH_TIME_DELTA - START_LUNCH_TIME_DELTA)
   
    return math.floor(time_available_in_day_time_delta.seconds * num_doctors / APPOINTMENT_LENGTH_TIME_DELTA.seconds)


#### VVVV BELOW IS UNIMPLEMENTED/BRAINSTORMING ONLY VVVV ####

def get_number_of_appointments_available_today_per_doctor(session, employee_id):
    pass
    
    time_available_in_day_time_delta = CLOSING_HOUR_TIME_DELTA - OPENING_HOUR_TIME_DELTA - \
                                       (END_LUNCH_TIME_DELTA - START_LUNCH_TIME_DELTA)
    
    doctor = session.query(Employee).filter(Employee.id == employee_id).first()
    
    
    return math.floor(time_available_in_day_time_delta.seconds * len(doctor) / APPOINTMENT_LENGTH_TIME_DELTA.seconds)

def check_if_speciality_doctor_is_available(session, appointment_datetime):
    pass
    
    if appointment_datetime <= get_todays_opening_datetime() or appointment_datetime >= get_todays_closing_datetime():
        return False
    if appointment_datetime >= START_LUNCH_TIME_DELTA and appointment_datetime < END_LUNCH_TIME_DELTA:
        return False
    
    todays_appointments = session.query(Appointment).filter(Appointment.date.between(get_todays_opening_datetime(), get_todays_closing_datetime()))
    
    if todays_appointments is None:
        return True
    
    doctors_with_appointments_today_dict = {}
    
    for appointment in todays_appointments:
        if appointment.doctor_id not in doctors_with_appointments_today_dict:
            doctors_with_appointments_today_dict[appointment.doctor_id] = []
        doctors_with_appointments_today_dict[appointment.doctor_id].append(appointment.datetime)
        
    # count up number of available appointments for each doctor
    # check if appointment_datetime is in the list of appointments for that doctor
    # if it is, then the doctor is not available
    # if it is not, then the doctor is available
        
    return True